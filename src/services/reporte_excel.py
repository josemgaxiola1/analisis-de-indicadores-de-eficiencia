import pandas as pd

from src.services.data_loader import cargar_todos_los_meses, clasificar_sucursales, MESES

COLOR_CATEGORIA = {
    "Excelente": "C6EFCE",
    "Cumple": "E2EFDA",
    "Más o menos": "FFEB9C",
    "Revisión": "FFC7CE",
}


def generar_excel(data_dir: str = "data/raw", salida: str = "reports/Analisis_Metas_Enero_Junio.xlsx") -> str:
    df = cargar_todos_los_meses(data_dir)
    clasif = clasificar_sucursales(df)
    revision = clasif[clasif["Categoría"] == "Revisión"].sort_values("Promedio_Estrellas")

    resumen_general = pd.DataFrame({
        "Indicador": [
            "Sucursales analizadas",
            "Periodo",
            "Promedio de estrellas (cadena)",
            "Sucursales en Revisión",
            "En Revisión y empeorando",
            "En Revisión y mejorando",
            "Mejor sucursal",
            "Peor sucursal",
        ],
        "Valor": [
            df["Sucursal"].nunique(),
            "Enero - Junio 2026",
            round(df["Estrellas Alcanzadas"].mean(), 2),
            len(revision),
            (revision["Tendencia"] == "📉 Empeorando").sum(),
            (revision["Tendencia"] == "📈 Mejorando").sum(),
            clasif.sort_values("Promedio_Estrellas", ascending=False).iloc[0]["Sucursal"],
            clasif.sort_values("Promedio_Estrellas").iloc[0]["Sucursal"],
        ],
    })

    evolucion_mensual = df.pivot_table(index="Sucursal", columns="Mes", values="Estrellas Alcanzadas")
    evolucion_mensual = evolucion_mensual[[m for m in MESES if m in evolucion_mensual.columns]]

    revision_detalle = revision.merge(
        evolucion_mensual, left_on="Sucursal", right_index=True
    )[["Sucursal", "Promedio_Estrellas", "Tendencia", "Métrica Débil",
       "Promedio_Ventas", "Promedio_Cantidad", "Promedio_Rentabilidad"] + MESES]

    clasif_export = clasif[[
        "Sucursal", "Categoría", "Promedio_Estrellas", "Tendencia", "Métrica Débil",
        "Promedio_Ventas", "Promedio_Cantidad", "Promedio_Rentabilidad",
    ]].sort_values(["Categoría", "Promedio_Estrellas"])

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        resumen_general.to_excel(writer, sheet_name="Resumen Ejecutivo", index=False)
        revision_detalle.to_excel(writer, sheet_name="Sucursales en Revision", index=False)
        clasif_export.to_excel(writer, sheet_name="Clasificacion Completa", index=False)
        df.to_excel(writer, sheet_name="Datos Detalle (6 meses)", index=False)

        _formatear(writer, "Resumen Ejecutivo", resumen_general)
        _formatear(writer, "Sucursales en Revision", revision_detalle, colorear_categoria=False)
        _formatear(writer, "Clasificacion Completa", clasif_export, colorear_categoria=True)
        _formatear(writer, "Datos Detalle (6 meses)", df, colorear_categoria=False)

    return salida


def _formatear(writer, hoja, df, colorear_categoria=False):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[hoja]
    header_fill = PatternFill("solid", fgColor="2A78D6")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for col_idx, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"

    if colorear_categoria and "Categoría" in df.columns:
        col_idx = df.columns.get_loc("Categoría") + 1
        for row_idx, valor in enumerate(df["Categoría"], start=2):
            color = COLOR_CATEGORIA.get(valor)
            if color:
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor=color)


if __name__ == "__main__":
    ruta = generar_excel()
    print(f"Excel generado en: {ruta}")
    ruta_kelder = generar_excel(
        data_dir="data/raw/KELDER", salida="reports/Analisis_Metas_Kelder_Enero_Junio.xlsx"
    )
    print(f"Excel generado en: {ruta_kelder}")
